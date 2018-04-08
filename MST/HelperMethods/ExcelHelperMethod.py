import csv
import xlsxwriter
from xlrd import open_workbook
from xlutils.copy import copy
import WordLetterProcessingHelperMethod
from openpyxl import Workbook, load_workbook

extension = 'csv'

diacritization_error_excel_file_path = \
    "D:\Repos\\TensorFlowRepo\\tfMST\\Book1.xlsx"

diacritization_error_without_last_letter_excel_file_path = \
    "D:\Repos\\TensorFlowRepo\\tfMST\\Book2.xls "

workbook = xlsxwriter.Workbook(diacritization_error_excel_file_path)
worksheet = workbook.add_worksheet()
worksheet.write(0, 0, 'actual letter')
worksheet.write(0, 1, 'expected letter')
worksheet.write(0, 2, 'undiac letter')
worksheet.write(0, 3, 'char location in word')
worksheet.write(0, 4, 'char location in sentence')
worksheet.write(0, 5, 'actual word')
worksheet.write(0, 6, 'expected word')
worksheet.write(0, 7, 'sentence number')
worksheet.write(0, 8, 'sentence')
worksheet.write(0, 9, 'actual diacritics')
worksheet.write(0, 10, 'expected diacritics')
worksheet.write(0, 11, 'actual diacritics english')
worksheet.write(0, 12, 'expected diacritics english')

workbook.close()

workbook2 = xlsxwriter.Workbook(diacritization_error_without_last_letter_excel_file_path)
worksheet2 = workbook2.add_worksheet()
worksheet2.write(0, 0, 'RNN OP')
worksheet2.write(0, 1, 'Required Target')
worksheet2.write(0, 2, 'Error Location')
worksheet2.write(0, 3, 'word contain error')
worksheet2.write(0, 4, 'location')
worksheet2.write(0, 5, 'sentence')
worksheet2.write(0, 6, 'value')
worksheet2.write(0, 10, 'expected diacritics english')
worksheet2.write(0, 11, 'actual diacritics english')
workbook2.close()

list1 = ['ُ', 'َ', 'ِ', 'ُّ', 'َّ', 'ِّ', 'ٌ', 'ً', 'ٍ', 'ٌّ', 'ًّ', 'ٍّ', '']
list2 = ['damma', 'fatha', 'kasra', 'sh+damma', 'sh+fatha', 'sh+kasra', 'tan+dam', 'tan+fat', 'tan+kas',
         'sh+tan+da', 'sh+tan+fa', 'sh+tan+kas', 'none']

dic = dict(zip(list1, list2))


def read_rnn_op_csv_file(csv_complete_path):

    rnn_op = []

    with open(csv_complete_path, 'rt') as csvfile:
        reader = csv.reader(csvfile, delimiter=';', quotechar='|')

        for row in reader:
            rnn_op.append(list(map(float, row)))

        return rnn_op


def write_data_into_excel_file(errors, current_sentence, current_row_in_excel_file, sentence_number):
        wb = open_workbook(diacritization_error_excel_file_path)
        w = copy(wb)
        worksheet = w.get_sheet(0)

        all_sentence = ''
        for each_word in current_sentence:
            all_sentence += each_word + ' '

        current_row_in_excel_file += 1
        column = 0

        for each_object in errors:
            worksheet.write(current_row_in_excel_file, column, each_object.actual_letter.letter)

            column = 1
            worksheet.write(current_row_in_excel_file, column, each_object.expected_letter.letter)

            column = 2
            worksheet.write(current_row_in_excel_file, column, each_object.error_location)

            column = 3
            worksheet.write(current_row_in_excel_file, column, each_object.exp_word)

            column = 4
            worksheet.write(current_row_in_excel_file, column, each_object.actual_letter.location)

            column = 5
            worksheet.write(current_row_in_excel_file, column, all_sentence)

            column = 6
            worksheet.write(current_row_in_excel_file, column, sentence_number)

            column = 7
            worksheet.write(current_row_in_excel_file, column, each_object.letter)

            column = 8
            worksheet.write(current_row_in_excel_file, column, each_object.expected_diacritics)

            column = 9
            worksheet.write(current_row_in_excel_file, column, each_object.actual_diacritics)

            column = 10
            worksheet.write(current_row_in_excel_file, column, dic[each_object.expected_diacritics])

            column = 11
            worksheet.write(current_row_in_excel_file, column, dic[each_object.actual_diacritics])

            current_row_in_excel_file += 1
            column = 0

        current_row_in_excel_file += 1

        w.save(diacritization_error_excel_file_path)
        workbook.close()

        return current_row_in_excel_file


def write_data_into_excel_file_version_2(errors):


    wb = load_workbook(diacritization_error_excel_file_path)
    #wb = open_workbook(diacritization_error_excel_file_path)
    #w = copy(wb)
    # worksheet = w.get_sheet(0)
    #wb = Workbook(write_only=True)

    worksheet1 = wb.worksheets[0]

    current_row_in_excel_file = 2

    for each_object in errors:
        column = 1
        #worksheet1.write(current_row_in_excel_file, column, each_object.actual_letter)
        worksheet1.cell(row = current_row_in_excel_file, column=column).value = each_object.actual_letter

        column += 1
        #worksheet1.write(current_row_in_excel_file, column, each_object.expected_letter)
        worksheet1.cell(row=current_row_in_excel_file, column=column).value = each_object.expected_letter

        column += 1
        #worksheet1.write(current_row_in_excel_file, column, each_object.undiac_letter)
        worksheet1.cell(row=current_row_in_excel_file, column=column).value = each_object.undiac_letter

        column += 1
        #worksheet1.write(current_row_in_excel_file, column, each_object.error_location_in_word)
        worksheet1.cell(row=current_row_in_excel_file, column=column).value = each_object.error_location_in_word

        column += 1
        #worksheet1.write(current_row_in_excel_file, column, each_object.error_location_in_sentence)
        worksheet1.cell(row=current_row_in_excel_file, column=column).value = each_object.error_location_in_sentence

        column += 1
        #worksheet1.write(current_row_in_excel_file, column, each_object.act_word)
        worksheet1.cell(row=current_row_in_excel_file, column=column).value = each_object.act_word

        column += 1
        #worksheet1.write(current_row_in_excel_file, column, each_object.exp_word)
        worksheet1.cell(row=current_row_in_excel_file, column=column).value = each_object.exp_word

        column += 1
        #worksheet1.write(current_row_in_excel_file, column, each_object.sentence_number)
        worksheet1.cell(row=current_row_in_excel_file, column=column).value = each_object.sentence_number

        column += 1
        #worksheet.write(current_row_in_excel_file, column, each_object.sentence)

        worksheet1.cell(row=current_row_in_excel_file, column=column).value = " ".join(each_object.sentence)

        column += 1
        #worksheet1.write(current_row_in_excel_file, column, each_object.actual_diacritics)
        worksheet1.cell(row=current_row_in_excel_file, column=column).value = each_object.actual_diacritics

        column += 1
        #worksheet1.write(current_row_in_excel_file, column, each_object.expected_diacritics)
        worksheet1.cell(row=current_row_in_excel_file, column=column).value = each_object.expected_diacritics

        column += 1
        #worksheet1.write(current_row_in_excel_file, column, dic[each_object.actual_diacritics])
        worksheet1.cell(row=current_row_in_excel_file, column=column).value = dic.get(each_object.actual_diacritics, 'empty')

        column += 1
        try:

            #worksheet1.write(current_row_in_excel_file, column, dic[each_object.expected_diacritics])
            worksheet1.cell(row=current_row_in_excel_file, column=column).value = dic.get(each_object.expected_diacritics, 'empty')

        except:
            x = 1
        current_row_in_excel_file += 1

    #w.save(diacritization_error_excel_file_path)
    wb.save('new_filename.xlsx')
    workbook.close()


def write_data_into_excel_file2(errors, current_sentence, current_row_in_excel_file, sentence_number):
    wb = open_workbook(diacritization_error_without_last_letter_excel_file_path)
    w = copy(wb)
    worksheet2 = w.get_sheet(0)

    all_sentence = ''
    for each_word in current_sentence:
        all_sentence += each_word + ' '

    current_row_in_excel_file += 1
    column = 0

    for each_object in errors:
        worksheet2.write(current_row_in_excel_file, column, each_object.actual_letter.letter)

        column = 1
        worksheet2.write(current_row_in_excel_file, column, each_object.expected_letter.letter)

        column = 2
        worksheet2.write(current_row_in_excel_file, column, each_object.error_location)

        column = 3
        worksheet2.write(current_row_in_excel_file, column, each_object.word)

        column = 4
        worksheet2.write(current_row_in_excel_file, column, each_object.actual_letter.location)

        column = 5
        worksheet2.write(current_row_in_excel_file, column, all_sentence)

        column = 6
        worksheet2.write(current_row_in_excel_file, column, sentence_number)

        current_row_in_excel_file += 1
        column = 0

    current_row_in_excel_file += 1

    w.save(diacritization_error_without_last_letter_excel_file_path)
    workbook.close()

    return current_row_in_excel_file


def write_data_into_excel_file2_version_2(errors):
    wb = open_workbook(diacritization_error_without_last_letter_excel_file_path)
    w = copy(wb)
    worksheet2 = w.get_sheet(0)

    current_row_in_excel_file = 1
    column = 0

    for each_object in errors:
        worksheet2.write(current_row_in_excel_file, column, each_object.actual_letter.letter)

        column = 1
        worksheet2.write(current_row_in_excel_file, column, each_object.expected_letter.letter)

        column = 2
        worksheet2.write(current_row_in_excel_file, column, each_object.error_location)

        column = 3
        worksheet2.write(current_row_in_excel_file, column, each_object.word)

        column = 4
        worksheet2.write(current_row_in_excel_file, column, each_object.actual_letter.location)

        column = 5
        worksheet2.write(current_row_in_excel_file, column, each_object.sentence)

        column = 6
        worksheet2.write(current_row_in_excel_file, column, each_object.sentence_number)

        current_row_in_excel_file += 1
        column = 0

    w.save(diacritization_error_without_last_letter_excel_file_path)
    workbook.close()


def read_csv_file(csv_complete_path):
    csv_op = []

    with open(csv_complete_path, 'rb') as csvfile:
        reader = csv.reader(csvfile, delimiter=';', quotechar='|')

        for row in reader:
            csv_op.append(unicode(row[0], "utf-8"))

        return csv_op